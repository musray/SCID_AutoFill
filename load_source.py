import os
import re
from openpyxl import load_workbook

def split_by_lang(message):
    # TODO 根据\n的位置，把message里的中文和英文分开
    # 规则：
    # 1. 连续英文、数字和空格
    # 2. 以`\n`或^开头
    # 3. 以`\n`或&结尾

    # 应对后英文先中文的情况
    pattern1 = r'(\n[a-zA-Z0-9 \n]+)$'
    # 应对先英文后中文的情况
    pattern2 = r'(^[a-zA-Z0-9 \n]+)\n'

    r1 = re.findall(pattern1, message)
    r2 = re.findall(pattern2, message)

    if r1:
        eng_message = r1[0]
        zh_message = re.sub(eng_message, '', message)
    elif r2:
        eng_message = r2[0]
        zh_message = re.sub(eng_message, '', message)
    else:
        eng_message = None
        zh_message = None

    return (eng_message, zh_message)

def load_source():

    files = os.listdir('./source')
    col_keyword = '功能描述'

    # 一会儿用来存放所有xlsx文件数据的池子
    pool = []

    for file in files:
        wb = load_workbook(os.path.join(os.getcwd(), 'source', file)) 
        ws = wb.active
        # id_des = [(row[3].value, [row[3].value]) for row in ws]
        id_des = []
        double_description_col = col_keyword in ws['D1'].value
        for row in ws.rows:
            # print(row)
            id = row[0].value
            if double_description_col:
                description = (row[2].value, row[3].value)
            else:
                description = split_by_lang(row[2].value)
            id_des.append((id, description))
        pool.extend(id_des)    
        wb.close()

    return pool